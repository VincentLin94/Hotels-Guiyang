# -*- coding:utf-8 -*-

from bs4 import BeautifulSoup
import requests
import re
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import time
import json
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet import worksheet


# First manual run to get the cookies
'''
browser = webdriver.Chrome()
browser.get('https://www.ctrip.com/')
# time.sleep(30)

# cookies = browser.get_cookies()
# print(type(cookies))
# print(cookies)
# f = open('cookie.txt', 'w')
# f.write(json.dumps(cookies))
# f.close()

f1 = open('cookie.txt')
cookie = f1.read()
cookie = json.loads(cookie)
for c in cookie:
    browser.add_cookie(c)
browser.refresh()

location = browser.find_element_by_xpath('//*[@id="HD_CityName"]')
location.clear()
location.send_keys('贵阳')
date_in = browser.find_element_by_xpath('//*[@id="HD_CheckIn"]')
date_in.clear()
date_in.send_keys('2020-07-26')
data_out = browser.find_element_by_xpath('//*[@id="HD_CheckOut"]')
data_out.clear()
data_out.send_keys('2020-07-27')
data_out.send_keys(Keys.RETURN)
'''


# Trial parsing the page source
'''
f = open('content.txt', 'r', encoding='utf-8')
content = f.read()
soup = BeautifulSoup(content, 'html.parser')
# print(soup)
hotels = soup.find_all('div', class_='hotel_new_list J_HotelListBaseCell')
print(hotels)
f.close()
'''
'''
# name, url, addrs, tags, facilities, price, score, comments, briefs = \
#     [], [], [], [], [], [], [], [], []
hotel_list = []
for i in hotels:
    hotel = i.find('a', href=True)
    name = hotel['title']
    url = 'https://hotels.ctrip.com' + hotel['href']
    url = url.replace('?isFull=F', '')
    addrs = i.find('p', class_='hotel_item_htladdress').text
    tags = i.find('span', class_='special_label').text
    icons = i.find('div', class_='icon_list').find_all('i')
    facilities = str()
    for f in icons:
        text = f['title']
        facilities = facilities + text
    price = i.find('span', class_='J_price_lowList').text
    score = i.find('a', class_='hotel_judge')['title']
    comments = i.find('span', class_='hotel_judgement').text
    briefs = i.find('span', class_='recommend')
    if briefs is None:
        briefs = 'None'
    else:
        briefs = briefs.text
    data = (name, url, addrs, tags, facilities, price, score, comments, briefs)
    hotel_list.append(data)

# print(hotel_list)
'''


# Writing the document using openpyxl
'''
wb = Workbook()
ws = wb.active
wb.save('Ctrip-Hotels.xlsx')
wb1 = openpyxl.load_workbook('Ctrip-Hotels.xlsx')
sheet = wb.create_sheet('Hotels by Ctrip', 0)
for i, data in enumerate(hotel_list):
    for j, value in enumerate(data):
        sheet.cell(row=i+1, column=j+1, value=value)

wb.save('Ctrip-Hotels.xlsx')
'''


# Test for ergodicity of tags in a list and getting its value by attr
'''
r = requests.get('http://www.okzy.co/')
r = r.text
bs = BeautifulSoup(r, 'html.parser')
li = bs.find_all('ul')
# print(li)
titles = []
hrefs = []
for i in li:
    tag = i.find('a', href=re.compile(r'\d{5}.html'))
    if tag is not None:
        titles.append(tag.text.strip())
        hrefs.append(tag['href'])
print(titles)
print(hrefs)

'''
