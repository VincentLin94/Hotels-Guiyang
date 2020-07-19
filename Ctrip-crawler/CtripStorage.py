# -*- coding:utf-8 -*-

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class Storage(object):

    def __init__(self):
        self.wb = openpyxl.load_workbook('Ctrip-Hotels.xlsx')
        self.sheet = self.wb.create_sheet('Hotel-list', 0)

    def write_head(self):
        '''
        Write the first row of the sheet
        :return:
        '''
        values = ['HotelName', 'RelevantURL', 'Address',
                  'Labels', 'Facilities', 'LowestPrice',
                  'Overview', 'ViewerCount', 'BriefComments']
        for i, value in enumerate(values):
            self.sheet.cell(row=1, column=i+1, value=value)

    def write_data(self, datas):
        '''
        Write the hotel info into the same sheet
        :param datas: Hotel info
        :return:
        '''
        for i, data in enumerate(datas):
            for j, value in enumerate(data):
                self.sheet.cell(row=i+2, column=j+1, value=value)
            print(str('%s threads of hotel info have been stored' % (i+1)))

    def write_end(self):
        '''
        Save the sheet at certain filepath
        :return:
        '''
        self.wb.save('Ctrip-Hotels.xlsx')




